"""app.reporting.export — CSV/Excel/PDF generation from a ReportResult.
PDF export needs a real QApplication (QTextDocument/QPrinter are Qt
objects), same requirement as the worker tests — see
tests/workers/test_base_worker.py's qapp fixture, mirrored here.
"""
import csv
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

import pytest

try:
    from PySide6.QtWidgets import QApplication
except ImportError:
    pytest.skip("PySide6 not available", allow_module_level=True)

from openpyxl import load_workbook

from app.reporting.export import (
    export_csv,
    export_excel,
    export_pdf,
    render_html_table,
    to_rows,
)
from app.schemas.reporting import ReportResult


@pytest.fixture(scope="module")
def qapp():
    try:
        app = QApplication.instance() or QApplication([])
    except Exception as exc:  # noqa: BLE001 - e.g. no display available
        pytest.skip(f"cannot create QApplication: {exc}")
    return app


def _sample_result() -> ReportResult:
    return ReportResult(
        title="Sample Report", generated_at=datetime(2026, 1, 15, 10, 30, tzinfo=timezone.utc),
        columns=["SKU", "Product", "Quantity", "Value", "As Of"],
        rows=[
            {"SKU": "A-1", "Product": "Widget", "Quantity": Decimal("12.500"),
             "Value": Decimal("125.75"), "As Of": date(2026, 1, 1)},
            {"SKU": "B-2", "Product": "Gadget & <Co>", "Quantity": Decimal("0"),
             "Value": Decimal("0"), "As Of": date(2026, 1, 2)},
        ])


def test_to_rows_converts_decimal_to_float_and_keeps_column_order():
    """Numeric, not text: a user has to be able to sum a Value column in
    Excel. The Decimal -> float conversion is display-only; the arithmetic
    that produced these numbers already happened in Decimal upstream."""
    rows = to_rows(_sample_result())

    assert rows[0] == ["A-1", "Widget", 12.5, 125.75, date(2026, 1, 1)]
    assert isinstance(rows[0][2], float)


def test_to_rows_of_an_empty_result_is_empty():
    empty = ReportResult(title="Empty", generated_at=datetime.now(timezone.utc),
                         columns=["A", "B"], rows=[])

    assert to_rows(empty) == []


def test_to_rows_fills_a_column_a_row_is_missing():
    """Reports assemble rows per query; an absent optional column should be
    a blank cell, not a failed export."""
    partial = ReportResult(title="Partial", generated_at=datetime.now(timezone.utc),
                           columns=["A", "B"], rows=[{"A": 1}])

    assert to_rows(partial) == [[1, None]]


def test_export_csv_writes_readable_rows(tmp_path):
    path = str(tmp_path / "report.csv")
    export_csv(_sample_result(), path)

    assert Path(path).exists()
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 2
    assert rows[0]["SKU"] == "A-1"
    assert rows[0]["Quantity"] == "12.5"


def test_export_csv_is_written_with_a_bom_so_excel_reads_utf8(tmp_path):
    """Without the BOM, Excel on Windows opens a UTF-8 CSV as the local
    code page and mangles every non-ASCII product name."""
    path = str(tmp_path / "report.csv")
    export_csv(_sample_result(), path)

    assert Path(path).read_bytes().startswith(b"\xef\xbb\xbf")


def test_export_csv_neutralizes_formula_injection(tmp_path):
    """CWE-1236: report cells echo free-text user input, and Excel executes
    a cell that starts with '=' on open."""
    hostile = _sample_result().model_copy(update={
        "rows": [{"SKU": "=cmd|'/c calc'!A1", "Product": "x", "Quantity": Decimal("1"),
                  "Value": Decimal("1"), "As Of": date(2026, 1, 1)}]})
    path = str(tmp_path / "report.csv")
    export_csv(hostile, path)

    with open(path, newline="", encoding="utf-8-sig") as f:
        row = next(csv.DictReader(f))
    assert row["SKU"].startswith("'")


def test_export_excel_writes_a_header_and_numeric_cells(tmp_path):
    path = str(tmp_path / "report.xlsx")
    export_excel(_sample_result(), path)

    assert Path(path).exists()
    sheet = load_workbook(path).active
    assert [cell.value for cell in sheet[1]] == \
        ["SKU", "Product", "Quantity", "Value", "As Of"]
    assert [row[0].value for row in sheet.iter_rows(min_row=2)] == ["A-1", "B-2"]
    # Numeric, not the string "125.75" — the user has to be able to sum it.
    assert sheet.cell(row=2, column=4).value == 125.75
    assert isinstance(sheet.cell(row=2, column=4).value, float)


def test_export_excel_sheet_name_truncated_to_31_chars(tmp_path):
    long_title = "A" * 50
    result = _sample_result().model_copy(update={"title": long_title})
    path = str(tmp_path / "report.xlsx")
    export_excel(result, path)

    assert len(load_workbook(path).sheetnames[0]) <= 31


def test_export_pdf_produces_a_nonempty_file(qapp, tmp_path):
    path = str(tmp_path / "report.pdf")
    export_pdf(_sample_result(), path)

    assert Path(path).exists()
    assert Path(path).stat().st_size > 0
    with open(path, "rb") as f:
        assert f.read(4) == b"%PDF"


def test_render_html_table_escapes_special_characters():
    html = render_html_table(_sample_result())
    assert "Gadget &amp; &lt;Co&gt;" in html
    assert "<Co>" not in html  # never appears unescaped
    assert "Sample Report" in html
    assert "2 row(s)" in html


def test_render_html_table_handles_empty_rows():
    empty = ReportResult(title="Empty", generated_at=datetime.now(timezone.utc),
                         columns=["A"], rows=[])
    html = render_html_table(empty)
    assert "<table>" in html
    assert "0 row(s)" in html
