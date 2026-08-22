"""
Tests for the storage layer.

Every test runs against a throwaway data folder. ``storage`` is imported at
module scope on purpose: resolving the data folder is lazy, so importing it
must NOT touch the filesystem — if that ever regresses, a test run would bind
to and mutate the developer's real workbooks.

Run with:  pytest -q
"""

import glob
import json
import os
import subprocess
import sys

import pytest
from openpyxl import Workbook, load_workbook

import storage


# The ``store`` fixture lives in conftest.py: it is autouse, so every test in
# this file is sandboxed whether or not it asks for the fixture by name.


def bill(store, is_sale=True, product="Widget", qty=1, rate=100, bill_no="B1",
         pan="AAAAA1111A", vendor="Acme", address="Addr", ecs=0, vat=13):
    """Commit a simple one-line bill and return its bill id."""
    lines = [{"product": product, "qty": qty, "rate": rate,
              "amount": store.line_amount(qty, rate)}]
    totals = store.bill_totals(lines, ecs, vat)
    header = {"date": "01/01/2026", "bill": bill_no, "pan": pan,
              "vendor": vendor, "address": address}
    return store.commit_bill(is_sale, header, lines, totals)["bill_id"]


def sales_path(store):
    return os.path.join(store.data_dir(), "sales.xlsx")


def stock_path(store):
    return os.path.join(store.data_dir(), "stock.xlsx")


def party_path(store):
    return os.path.join(store.data_dir(), "party.xlsx")


def crash_swapping(monkeypatch, filename):
    """Make the next os.replace ONTO ``filename`` fail once, as a crash would.

    Targeted by destination rather than by call count, so unrelated writes
    (creating a workbook, refreshing a .bak) cannot consume the trigger.
    Returns the real os.replace so the caller can restore it.
    """
    real_replace = os.replace
    state = {"armed": True}

    def flaky(src, dst):
        if state["armed"] and os.path.basename(str(dst)) == filename:
            state["armed"] = False
            raise OSError("simulated crash between swaps")
        return real_replace(src, dst)

    monkeypatch.setattr(os, "replace", flaky)
    return real_replace


# --------------------------------------------------------------------------- #
# Import hygiene
# --------------------------------------------------------------------------- #
def test_importing_storage_does_not_touch_the_filesystem(tmp_path):
    """The data folder must be resolved lazily, never at import time.

    Checked in a subprocess, because by the time this process runs a test the
    module has long since been imported and the property is unobservable here.
    """
    target = tmp_path / "untouched"
    env = dict(os.environ, INVENTORY_DATA_DIR=str(target))
    proc = subprocess.run(
        [sys.executable, "-c", "import storage; print('imported')"],
        cwd=os.path.dirname(os.path.abspath(storage.__file__)),
        env=env, capture_output=True, text=True)
    assert proc.returncode == 0, proc.stderr
    assert not target.exists(), (
        f"importing storage created {target} -- the data folder is being "
        f"resolved at import time again")


def test_path_attributes_resolve_lazily(store):
    assert store.SALES_FILE == os.path.join(store.data_dir(), "sales.xlsx")
    assert store.STOCK_FILE == os.path.join(store.data_dir(), "stock.xlsx")
    with pytest.raises(AttributeError):
        _ = store.NOT_A_REAL_ATTRIBUTE


def test_data_dir_expands_user_and_vars(store, tmp_path, monkeypatch):
    """The override must expand $VARS and ~ rather than making a literal folder.

    Kept inside the sandbox (tmp_path/data) so the conftest guard still holds.
    """
    monkeypatch.setenv("MY_BOOKS", str(tmp_path / "data"))
    monkeypatch.setenv("INVENTORY_DATA_DIR", "$MY_BOOKS")
    storage.reset_data_dir()
    resolved = storage.data_dir()
    assert "$MY_BOOKS" not in resolved
    assert os.path.realpath(resolved).startswith(
        os.path.realpath(str(tmp_path / "data")))
    assert resolved.endswith("inventory_data"), "override is a BASE folder"


# --------------------------------------------------------------------------- #
# Value parsing
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("raw,expected", [
    (None, 0.0), ("", 0.0), ("abc", 0.0), ("13%", 13.0), ("1,030", 1030.0),
    ("  42 ", 42.0), ("-5", -5.0), ("1e400", 0.0), (7, 7.0),
])
def test_num_parses_form_input(raw, expected):
    assert storage.num(raw) == expected


def test_stored_num_refuses_junk_instead_of_defaulting_to_zero():
    """A ledger balance that will not parse is an error, never a silent 0."""
    assert storage._stored_num(None, "f.xlsx", 2, "Quantity") == 0.0
    assert storage._stored_num("", "f.xlsx", 2, "Quantity") == 0.0
    assert storage._stored_num(5, "f.xlsx", 2, "Quantity") == 5.0
    assert storage._stored_num("1,030", "f.xlsx", 2, "Quantity") == 1030.0
    with pytest.raises(storage.DataIntegrityError):
        storage._stored_num("=1+1", "f.xlsx", 2, "Quantity")
    with pytest.raises(storage.DataIntegrityError):
        storage._stored_num("banana", "f.xlsx", 2, "Quantity")


def test_key_matches_sanitized_and_raw_forms():
    """The whole point: what we write and what we look up must compare equal."""
    for name in ("-Special", "+2 Lens", "=Total", "@Home", "Widget"):
        assert storage._key(storage._safe_text(name)) == storage._key(name)


def test_safe_text_still_neutralises_formulas():
    assert storage._safe_text("=SUM(A1:A9)").startswith("'")
    assert storage._safe_text("Widget") == "Widget"


# --------------------------------------------------------------------------- #
# Business rules
# --------------------------------------------------------------------------- #
def test_bill_totals_is_the_single_definition_of_the_math():
    lines = [{"amount": 100.0}, {"amount": 50.0}]
    t = storage.bill_totals(lines, ecs=10, vat_pct=13)
    assert t["subtotal"] == 150.0
    assert t["vat_amount"] == 19.5
    assert t["total"] == 179.5


def test_line_amount_rejects_unstorable_numbers():
    assert storage.line_amount(3, 12.5) == 37.5
    with pytest.raises(ValueError):
        storage.line_amount(1e308, 10)


def test_bill_totals_rejects_overflow():
    """Regression: inf silently became a blank money cell in the ledger."""
    with pytest.raises(ValueError):
        storage.bill_totals([{"amount": 9e11}, {"amount": 9e11}], 0, 13)
    ok = storage.bill_totals([{"amount": 1e6}], 0, 13)
    assert ok["total"] == 1130000.0


# --------------------------------------------------------------------------- #
# Stock
# --------------------------------------------------------------------------- #
def test_stock_accumulates_across_bills(store):
    bill(store, is_sale=False, qty=10, bill_no="P1")
    bill(store, is_sale=False, qty=5, bill_no="P2")
    assert store.stock_on_hand("Widget") == 15
    bill(store, is_sale=True, qty=4, bill_no="S1")
    assert store.stock_on_hand("Widget") == 11


def test_product_name_starting_with_a_formula_char_still_accumulates(store):
    """Regression: sanitize-on-write / match-on-raw forked a row per bill."""
    bill(store, is_sale=False, product="-Special", qty=10, bill_no="P1")
    bill(store, is_sale=False, product="-Special", qty=10, bill_no="P2")
    assert store.stock_on_hand("-Special") == 20

    ws = load_workbook(stock_path(store))[storage.SHEET_NAME]
    products = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert len(products) == 1, f"stock forked into duplicate rows: {products}"
    assert store.product_names() == ["-Special"], "dropdown shows the typed name"


def test_stock_matching_is_case_and_whitespace_insensitive(store):
    bill(store, is_sale=False, product="Widget", qty=5, bill_no="P1")
    bill(store, is_sale=False, product="  wIdGeT ", qty=5, bill_no="P2")
    assert store.stock_on_hand("widget") == 10


def test_shortages_sums_repeated_products_in_one_bill(store):
    """Regression: two lines of 5 against 8 on hand raised no warning."""
    bill(store, is_sale=False, product="Gadget", qty=8, bill_no="P1")
    lines = [{"product": "Gadget", "qty": 5, "rate": 1, "amount": 5},
             {"product": "Gadget", "qty": 5, "rate": 1, "amount": 5}]
    short = store.shortages(lines)
    assert len(short) == 1
    assert short[0]["have"] == 8
    assert short[0]["selling"] == 10


def test_shortages_is_quiet_when_stock_is_sufficient(store):
    bill(store, is_sale=False, product="Gadget", qty=20, bill_no="P1")
    lines = [{"product": "Gadget", "qty": 5, "rate": 1, "amount": 5},
             {"product": "Gadget", "qty": 5, "rate": 1, "amount": 5}]
    assert store.shortages(lines) == []


# --------------------------------------------------------------------------- #
# Party totals
# --------------------------------------------------------------------------- #
def test_party_totals_accumulate_and_stay_rounded(store):
    """Regression: repeated += on floats displayed 656.9900000000002.

    0.1 + 0.2 != 0.3 in binary floating point, so ten bills of 0.10 are the
    case that actually produces a drifting tail without the round().
    """
    for i in range(10):
        lines = [{"product": "W", "qty": 1, "rate": 0.1, "amount": 0.1}]
        totals = store.bill_totals(lines, 0, 0)
        store.commit_bill(True, {"date": "d", "bill": f"S{i}",
                                 "pan": "AAAAA1111A", "vendor": "Acme",
                                 "address": "A"}, lines, totals)
    ws = load_workbook(party_path(store))[storage.SHEET_NAME]
    assert ws.max_row == 2, "one party, one row"
    stored = ws.cell(row=2, column=storage.P_SALES + 1).value
    assert stored == 1.0, f"float tail leaked into the ledger: {stored!r}"
    assert ws.cell(row=2, column=storage.P_COMBINED + 1).value == 1.0


def test_vendor_name_starting_with_a_formula_char_does_not_fork(store):
    bill(store, pan="", vendor="-Traders", bill_no="S1")
    bill(store, pan="", vendor="-Traders", bill_no="S2")
    ws = load_workbook(party_path(store))[storage.SHEET_NAME]
    assert ws.max_row == 2, "party forked into duplicate rows"


def test_blank_pan_parties_are_kept_apart_by_address(store):
    bill(store, pan="", vendor="Sharma", address="Delhi", bill_no="S1")
    bill(store, pan="", vendor="Sharma", address="Mumbai", bill_no="S2")
    ws = load_workbook(party_path(store))[storage.SHEET_NAME]
    assert ws.max_row == 3, "two different vendors were merged"


# --------------------------------------------------------------------------- #
# Bill identity and grouping
# --------------------------------------------------------------------------- #
def test_two_blank_bill_no_bills_stay_separate(store):
    """Regression: identical header columns collapsed two bills into one."""
    bill(store, bill_no="", qty=1, rate=100)     # total 113
    bill(store, bill_no="", qty=2, rate=100)     # total 226
    found = store.bills(sales_path(store))
    assert len(found) == 2, "two distinct bills merged into one"
    assert round(sum(b["totals"]["total"] for b in found), 2) == 339.0


def test_two_bills_sharing_one_bill_no_stay_separate(store):
    """Grouping must key on Bill ID, not on the visible header columns.

    A shop that reuses a bill number (or types the same one twice) must still
    see two bills and a grand total that includes both.
    """
    bill(store, bill_no="B1", qty=1, rate=100)     # total 113
    bill(store, bill_no="B1", qty=2, rate=100)     # total 226
    found = store.bills(sales_path(store))
    assert len(found) == 2, "same Bill No collapsed two bills into one"
    assert {b["bill_id"] for b in found} == {1, 2}
    assert round(sum(b["totals"]["total"] for b in found), 2) == 339.0


def test_rows_without_a_bill_id_are_never_merged(store):
    """Regression: every id-less row collapsed into one pseudo-bill 0."""
    path = sales_path(store)
    bill(store, bill_no="B1", qty=1, rate=100)
    wb = load_workbook(path)
    ws = wb[storage.SHEET_NAME]
    for values, total in ((["01/01/2026", "X1", "P9", "Ann", "Addr", "W",
                            1, 10, 10, 10, 0, 0, 0, 10], 10),
                          (["02/01/2026", "X2", "P8", "Bob", "Addr", "W",
                            1, 20, 20, 20, 0, 0, 0, 20], 20)):
        ws.append(values)                      # no Bill ID column written
    wb.save(path)

    found = store.bills(path)
    assert len(found) == 3, "id-less rows were merged into one bill"
    assert sum(1 for b in found if b["synthetic"]) == 2
    assert round(sum(b["totals"]["total"] for b in found), 2) == 143.0

    # and they must not be voidable, since their line grouping is unknowable
    orphan = next(b for b in found if b["synthetic"])
    with pytest.raises(ValueError):
        store.void_bill(True, orphan["bill_id"])


def test_rebuild_party_totals_counts_id_less_rows(store):
    """Regression: dedupe on the raw Bill ID dropped every blank-ID row."""
    path = sales_path(store)
    bill(store, bill_no="B1", qty=1, rate=100, pan="PAN1")
    wb = load_workbook(path)
    ws = wb[storage.SHEET_NAME]
    ws.append(["01/01/2026", "X1", "PAN2", "Ann", "Addr", "W",
               1, 10, 10, 10, 0, 0, 0, 10])
    ws.append(["02/01/2026", "X2", "PAN3", "Bob", "Addr", "W",
               1, 20, 20, 20, 0, 0, 0, 20])
    wb.save(path)

    assert store.rebuild_party_totals() == 3, "an id-less party was dropped"
    rows = {r[storage.P_PAN]: r[storage.P_SALES]
            for r in store.read_rows(party_path(store), storage.PARTY_HEADERS)}
    assert rows == {"PAN1": 113.0, "PAN2": 10.0, "PAN3": 20.0}


def test_blank_bill_no_is_auto_assigned(store):
    bid = bill(store, bill_no="")
    found = store.bills(sales_path(store))[0]
    assert found["header"]["bill"] == f"AUTO-S{bid:05d}"


def test_grouping_survives_the_sheet_being_re_sorted(store):
    """The user is told they may open these files in Excel; sorting must be safe."""
    lines = [{"product": p, "qty": 1, "rate": 10, "amount": 10}
             for p in ("Zeta", "Alpha", "Mid")]
    totals = store.bill_totals(lines, 0, 13)
    store.commit_bill(True, {"date": "d", "bill": "B9", "pan": "P",
                             "vendor": "V", "address": "A"}, lines, totals)
    bill(store, bill_no="B10", qty=1, rate=50)

    before = store.bills(sales_path(store))
    grand_before = round(sum(b["totals"]["total"] for b in before), 2)

    # Simulate a user sorting the whole sheet by Product Name in Excel.
    wb = load_workbook(sales_path(store))
    ws = wb[storage.SHEET_NAME]
    data = [[ws.cell(row=r, column=c).value
             for c in range(1, len(storage.TXN_HEADERS) + 1)]
            for r in range(2, ws.max_row + 1)]
    data.sort(key=lambda row: str(row[storage.C_PRODUCT]))
    for r, row in enumerate(data, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    wb.save(sales_path(store))

    after = store.bills(sales_path(store))
    assert len(after) == len(before) == 2
    assert round(sum(b["totals"]["total"] for b in after), 2) == grand_before


def test_bill_exists_finds_a_formula_prefixed_bill_number(store):
    bill(store, bill_no="=BILL1")
    assert store.bill_exists(sales_path(store), "=BILL1") is True
    assert store.bill_exists(sales_path(store), "nope") is False
    assert store.bill_exists(sales_path(store), "") is False


# --------------------------------------------------------------------------- #
# Wrong-sheet protection
# --------------------------------------------------------------------------- #
def test_a_user_added_sheet_does_not_fork_the_ledger(store):
    """Regression: wb.active silently redirected reads and writes."""
    bill(store, is_sale=False, qty=50, bill_no="P1")
    wb = load_workbook(stock_path(store))
    wb.create_sheet("Notes")
    wb.active = wb["Notes"]
    wb.save(stock_path(store))

    assert store.stock_on_hand("Widget") == 50
    assert store.product_names() == ["Widget"]
    bill(store, is_sale=False, qty=5, bill_no="P2")
    assert store.stock_on_hand("Widget") == 55

    ws = load_workbook(stock_path(store))["Notes"]
    assert ws.max_row <= 1, "data was written into the user's own sheet"


# --------------------------------------------------------------------------- #
# Durability
# --------------------------------------------------------------------------- #
def test_a_failure_while_staging_changes_nothing(store, monkeypatch):
    """Regression: a mid-commit failure left the bill recorded but stock stale."""
    bill(store, is_sale=False, product="Widget", qty=100, bill_no="P1")
    bill(store, is_sale=True, product="Widget", qty=1, bill_no="S0")
    before = {p: open(p, "rb").read()
              for p in (sales_path(store), stock_path(store), party_path(store))}

    calls = {"n": 0}
    real_stage = storage._stage

    def flaky_stage(wb, dest):
        calls["n"] += 1
        if calls["n"] == 2:                      # fail on the stock workbook
            raise storage.StorageError("simulated disk full")
        return real_stage(wb, dest)

    monkeypatch.setattr(storage, "_stage", flaky_stage)
    with pytest.raises(storage.StorageError):
        bill(store, is_sale=True, product="Widget", qty=5, bill_no="S1")
    monkeypatch.setattr(storage, "_stage", real_stage)

    for path, content in before.items():
        assert open(path, "rb").read() == content, f"{path} was modified"
    assert store.stock_on_hand("Widget") == 99
    assert store.bill_exists(sales_path(store), "S1") is False
    assert not os.path.exists(storage._journal_path()), "a journal was left behind"


def test_an_interrupted_swap_is_finished_on_the_next_start(store, monkeypatch):
    """The journal is what makes a three-file commit effectively atomic.

    Note: no monkeypatch.undo() here. It would also revert the sandbox
    environment set up by the autouse fixture, pointing storage at the
    developer's real workbooks.
    """
    bill(store, is_sale=False, product="Widget", qty=100, bill_no="P1")

    real_replace = crash_swapping(monkeypatch, "stock.xlsx")
    with pytest.raises(storage.CommitInterruptedError):
        bill(store, is_sale=True, product="Widget", qty=5, bill_no="S1")
    monkeypatch.setattr(os, "replace", real_replace)

    assert os.path.exists(storage._journal_path()), "no journal to recover from"
    result = store.recover()
    assert result["recovered"] is True and result["failed"] == []
    assert not os.path.exists(storage._journal_path())
    assert store.bill_exists(sales_path(store), "S1") is True
    assert store.stock_on_hand("Widget") == 95, "caches were reconciled"


def test_recovery_reconciles_the_caches_even_if_their_swap_was_lost(store,
                                                                    monkeypatch):
    """recover()'s rebuild is load-bearing, not decoration."""
    bill(store, is_sale=False, product="Widget", qty=100, bill_no="P1")

    real_replace = crash_swapping(monkeypatch, "stock.xlsx")
    with pytest.raises(storage.CommitInterruptedError):
        bill(store, is_sale=True, product="Widget", qty=5, bill_no="S1")
    monkeypatch.setattr(os, "replace", real_replace)

    # Lose the staged stock workbook entirely, as a hard crash could.
    with open(storage._journal_path(), encoding="utf-8") as fh:
        pairs = json.load(fh)["pairs"]
    for tmp, dest in pairs:
        if dest.endswith("stock.xlsx") and os.path.exists(tmp):
            os.remove(tmp)

    result = store.recover()
    # A staged file that is listed but gone was LOST, and must be reported as
    # such rather than mistaken for a swap that already happened.
    assert result["recovered"] is False
    assert "lost" in result["message"].lower()
    assert "stock.xlsx" in result["message"]
    # The swap for stock.xlsx never happened, so only the rebuild can make the
    # cache agree with the ledger.
    assert store.stock_on_hand("Widget") == 95


def test_recover_keeps_the_journal_when_it_cannot_finish(store, monkeypatch):
    """A pending bill must never be abandoned silently."""
    bill(store, is_sale=False, product="Widget", qty=100, bill_no="P1")

    real_replace = crash_swapping(monkeypatch, "stock.xlsx")
    with pytest.raises(storage.CommitInterruptedError):
        bill(store, is_sale=True, product="Widget", qty=5, bill_no="S1")

    def always_fails(src, dst):
        if os.path.basename(str(dst)) == "stock.xlsx":
            raise OSError("still locked")
        return real_replace(src, dst)

    monkeypatch.setattr(os, "replace", always_fails)
    result = store.recover()
    monkeypatch.setattr(os, "replace", real_replace)

    assert result["failed"], "the failure was not reported"
    assert "restart" in result["message"].lower()
    assert os.path.exists(storage._journal_path()), (
        "the journal was thrown away, abandoning the pending bill")


def test_a_new_save_refuses_to_run_over_an_unfinished_one(store, monkeypatch):
    """Regression: the next save truncated the pending journal."""
    bill(store, is_sale=False, product="Widget", qty=100, bill_no="P1")

    real_replace = crash_swapping(monkeypatch, "stock.xlsx")
    with pytest.raises(storage.CommitInterruptedError):
        bill(store, is_sale=True, product="Widget", qty=5, bill_no="S1")
    monkeypatch.setattr(os, "replace", real_replace)

    with pytest.raises(storage.StorageError) as err:
        bill(store, is_sale=True, product="Widget", qty=1, bill_no="S2")
    assert "interrupted" in str(err.value).lower()
    assert os.path.exists(storage._journal_path()), "journal was destroyed"

    store.recover()                       # the documented way out
    bill(store, is_sale=True, product="Widget", qty=1, bill_no="S2")
    assert store.stock_on_hand("Widget") == 94


def test_recover_is_a_no_op_without_a_journal(store):
    bill(store, bill_no="S1")
    result = store.recover()
    assert result == {"recovered": False, "failed": [], "message": ""}


def test_orphaned_staging_files_are_swept_at_startup(store):
    bill(store, bill_no="S1")
    orphan = os.path.join(store.data_dir(), ".tmp-leftover.xlsx")
    with open(orphan, "wb") as fh:
        fh.write(b"junk")
    store.recover()
    assert not os.path.exists(orphan), "staging files accumulate forever"


def test_stage_turns_a_permission_error_into_a_clear_message(store):
    """The common real failure: the owner left the workbook open in Excel."""
    class LockedWorkbook:
        def save(self, path):
            raise PermissionError(13, "in use")

    with pytest.raises(storage.FileLockedError) as err:
        storage._stage(LockedWorkbook(), stock_path(store))
    assert "open in Excel/LibreOffice" in str(err.value)
    assert not glob.glob(os.path.join(store.data_dir(), ".tmp-*.xlsx")), \
        "the failed staging file was left behind"


def test_a_locked_workbook_leaves_the_books_untouched(store, monkeypatch):
    bill(store, is_sale=False, product="Widget", qty=10, bill_no="P1")
    real_stage = storage._stage

    def locked(wb, dest):
        if dest.endswith("stock.xlsx"):
            raise storage.FileLockedError(dest)
        return real_stage(wb, dest)

    monkeypatch.setattr(storage, "_stage", locked)
    with pytest.raises(storage.FileLockedError) as err:
        bill(store, is_sale=True, product="Widget", qty=1, bill_no="S1")
    assert "open in Excel/LibreOffice" in str(err.value)
    monkeypatch.setattr(storage, "_stage", real_stage)
    assert store.stock_on_hand("Widget") == 10, "nothing should have changed"
    assert store.bill_exists(sales_path(store), "S1") is False
    assert not os.path.exists(storage._journal_path())


def test_assert_writable_reports_a_locked_file(store, monkeypatch):
    bill(store, is_sale=False, product="Widget", qty=10, bill_no="P1")
    real_open = open

    def refuse(path, *a, **kw):
        if str(path).endswith("party.xlsx"):
            raise PermissionError(13, "in use")
        return real_open(path, *a, **kw)

    monkeypatch.setattr("builtins.open", refuse)
    with pytest.raises(storage.FileLockedError):
        store.assert_writable(party_path(store))


def test_every_save_keeps_the_previous_version(store, tmp_path):
    bill(store, is_sale=False, qty=10, bill_no="P1")
    bill(store, is_sale=False, qty=10, bill_no="P2")
    assert store.stock_on_hand("Widget") == 20

    backup = stock_path(store) + ".bak"
    assert os.path.exists(backup), "no .bak was kept"
    # openpyxl refuses the .bak suffix, so recover it the way a user would.
    recovered = tmp_path / "recovered.xlsx"
    recovered.write_bytes(open(backup, "rb").read())
    ws = load_workbook(recovered)[storage.SHEET_NAME]
    assert ws.cell(row=2, column=2).value == 10, "backup holds the prior state"


def test_daily_backup_copies_every_workbook(store):
    bill(store, bill_no="S1")
    folder = store.backup_daily()
    for name in ("sales.xlsx", "stock.xlsx", "party.xlsx"):
        assert os.path.exists(os.path.join(folder, name)), name


def test_a_damaged_workbook_reports_itself_clearly(store):
    bill(store, bill_no="S1")
    with open(sales_path(store), "wb") as fh:
        fh.write(b"not a zip file at all")
    with pytest.raises(storage.StorageError) as err:
        store.read_rows(sales_path(store), storage.TXN_HEADERS)
    assert "damaged" in str(err.value)


# --------------------------------------------------------------------------- #
# Rebuilds and void
# --------------------------------------------------------------------------- #
def test_rebuild_stock_repairs_a_tampered_cache(store):
    bill(store, is_sale=False, product="Widget", qty=30, bill_no="P1")
    bill(store, is_sale=True, product="Widget", qty=12, bill_no="S1")
    assert store.stock_on_hand("Widget") == 18

    wb = load_workbook(stock_path(store))          # someone edits it wrongly
    wb[storage.SHEET_NAME].cell(row=2, column=2, value=999)
    wb.save(stock_path(store))
    assert store.stock_on_hand("Widget") == 999

    store.rebuild_stock()
    assert store.stock_on_hand("Widget") == 18


def test_rebuild_party_totals_matches_the_ledgers(store):
    bill(store, is_sale=True, qty=1, rate=100, bill_no="S1")     # 113
    bill(store, is_sale=False, qty=1, rate=200, bill_no="P1")    # 226
    store.rebuild_party_totals()
    ws = load_workbook(party_path(store))[storage.SHEET_NAME]
    assert ws.cell(row=2, column=storage.P_SALES + 1).value == 113.0
    assert ws.cell(row=2, column=storage.P_PURCH + 1).value == 226.0
    assert ws.cell(row=2, column=storage.P_COMBINED + 1).value == 339.0


def test_rebuild_merges_ledger_rows_the_old_bug_had_spelled_two_ways(store):
    """Existing damaged files must be repairable, not just newly-safe.

    The old version wrote "'-Special" into the ledger for a product typed as
    "-Special", so a real damaged file has BOTH spellings. The rebuild must
    fold them into one product.
    """
    path = os.path.join(store.data_dir(), "purchases.xlsx")
    bill(store, is_sale=False, product="-Special", qty=10, bill_no="P1")
    wb = load_workbook(path)
    ws = wb[storage.SHEET_NAME]
    ws.append(["01/01/2026", "P2", "P", "V", "A", "'-Special",
               10, 1, 10, 10, 0, 0, 0, 10, 2, "", None])
    wb.save(path)

    assert store.rebuild_stock() == 1, "the two spellings were not folded"
    ws = load_workbook(stock_path(store))[storage.SHEET_NAME]
    assert ws.max_row == 2
    assert store.stock_on_hand("-Special") == 20


def test_stock_on_hand_sums_duplicate_rows_left_by_the_old_bug(store):
    """A damaged cache must read correctly even before it is rebuilt."""
    bill(store, is_sale=False, product="-Special", qty=10, bill_no="P1")
    wb = load_workbook(stock_path(store))          # forge the old broken state
    wb[storage.SHEET_NAME].append(["'-Special", 10])
    wb.save(stock_path(store))
    assert store.stock_on_hand("-Special") == 20, "half the stock was ignored"

    # and the next bill folds them back into one row
    bill(store, is_sale=False, product="-Special", qty=5, bill_no="P2")
    assert store.stock_on_hand("-Special") == 25
    ws = load_workbook(stock_path(store))[storage.SHEET_NAME]
    quantities = [ws.cell(row=r, column=2).value
                  for r in range(2, ws.max_row + 1)]
    assert sorted(quantities) == [0, 25], f"rows were not folded: {quantities}"


def test_void_bill_reverses_stock_and_party_totals(store):
    bill(store, is_sale=False, product="Widget", qty=50, bill_no="P1")
    sale_id = bill(store, is_sale=True, product="Widget", qty=10, rate=100,
                   bill_no="S1")
    assert store.stock_on_hand("Widget") == 40

    store.void_bill(True, sale_id)
    assert store.stock_on_hand("Widget") == 50, "stock was not restored"
    ws = load_workbook(party_path(store))[storage.SHEET_NAME]
    assert ws.cell(row=2, column=storage.P_SALES + 1).value == 0.0


def test_void_keeps_the_original_rows_for_the_audit_trail(store):
    sale_id = bill(store, bill_no="S1")
    store.void_bill(True, sale_id)
    found = store.bills(sales_path(store))
    assert len(found) == 2, "the original bill must not be deleted"
    original = next(b for b in found if b["bill_id"] == sale_id)
    assert original["voided_by"] is not None


def test_a_bill_cannot_be_voided_twice(store):
    sale_id = bill(store, bill_no="S1")
    store.void_bill(True, sale_id)
    with pytest.raises(ValueError):
        store.void_bill(True, sale_id)


def test_voiding_an_unknown_bill_is_an_error(store):
    bill(store, bill_no="S1")
    with pytest.raises(ValueError):
        store.void_bill(True, 9999)


# --------------------------------------------------------------------------- #
# Schema migration
# --------------------------------------------------------------------------- #
def test_a_legacy_14_column_ledger_is_upgraded_in_place(store):
    """Existing users' files must keep working and gain bill identity."""
    path = sales_path(store)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"                                   # old sheet name
    ws.append(storage._LEGACY_TXN_HEADERS)
    ws.append(["01/01/2026", "B1", "P1", "Acme", "Addr",
               "Widget", 2, 10, 20, 20, 0, 13, 2.6, 22.6])
    ws.append(["01/01/2026", "B1", "P1", "Acme", "Addr",
               "Bolt", 1, 30, 30, 20, 0, 13, 2.6, 22.6])
    ws.append(["02/01/2026", "B2", "P2", "Beta", "Addr2",
               "Nut", 1, 40, 40, 40, 0, 13, 5.2, 45.2])
    wb.save(path)

    found = store.bills(path)
    assert len(found) == 2, "historical bills were not reconstructed"
    assert len(found[0]["lines"]) == 2
    assert found[0]["header"]["bill"] == "B1"
    assert round(sum(b["totals"]["total"] for b in found), 2) == 67.8

    # A read migrates in memory only -- reads must never rewrite the user's
    # file. The upgrade is persisted by the next write.
    assert load_workbook(path).sheetnames == ["Sheet"]
    bill(store, bill_no="B3")
    wb = load_workbook(path)
    assert wb.sheetnames == [storage.SHEET_NAME], "sheet was not renamed on write"
    ws = wb[storage.SHEET_NAME]
    assert [ws.cell(row=1, column=c).value
            for c in range(1, len(storage.TXN_HEADERS) + 1)] == storage.TXN_HEADERS


def test_a_new_bill_after_migration_gets_a_fresh_id(store):
    path = sales_path(store)
    wb = Workbook()
    ws = wb.active
    ws.append(storage._LEGACY_TXN_HEADERS)
    ws.append(["01/01/2026", "B1", "P1", "Acme", "Addr",
               "Widget", 2, 10, 20, 20, 0, 13, 2.6, 22.6])
    wb.save(path)

    new_id = bill(store, bill_no="B2")
    assert new_id == 2
    assert len(store.bills(path)) == 2


def test_a_formula_in_a_managed_column_is_refused_not_read_as_zero(store):
    """Regression: an uncalculated formula read as 0 and wiped the balance."""
    bill(store, is_sale=False, product="Widget", qty=30, bill_no="P1")
    wb = load_workbook(stock_path(store))
    wb[storage.SHEET_NAME].cell(row=2, column=storage.S_QTY + 1, value="=15+15")
    wb.save(stock_path(store))

    with pytest.raises(storage.DataIntegrityError) as err:
        store.stock_on_hand("Widget")
    assert "Quantity" in str(err.value)

    with pytest.raises(storage.DataIntegrityError):
        bill(store, is_sale=False, product="Widget", qty=5, bill_no="P2")

    # and the repair path works
    store.rebuild_stock()
    assert store.stock_on_hand("Widget") == 30


def test_rebuild_survives_a_formula_left_in_the_stock_cache(store):
    bill(store, is_sale=False, product="Widget", qty=30, bill_no="P1")
    wb = load_workbook(stock_path(store))
    wb[storage.SHEET_NAME].cell(row=2, column=storage.S_QTY + 1, value="=oops")
    wb.save(stock_path(store))
    assert store.rebuild_stock() == 1
    assert store.stock_on_hand("Widget") == 30


# --------------------------------------------------------------------------- #
# Round-2 fixes
# --------------------------------------------------------------------------- #
def test_the_ledger_is_found_by_its_header_row_not_its_position(store):
    """Regression: worksheets[0] adopted the user's own tab and orphaned data."""
    path = sales_path(store)
    wb = Workbook()
    mine = wb.active
    mine.title = "My Notes"                    # user's tab, dragged to the front
    mine.append(["shopping", "list"])
    led = wb.create_sheet("Sheet")             # the real, pre-upgrade ledger
    led.append(storage._LEGACY_TXN_HEADERS)
    led.append(["01/01/2026", "B1", "P", "V", "A", "Widget",
                2, 10, 20, 20, 0, 13, 2.6, 22.6])
    wb.save(path)

    found = store.bills(path)
    assert len(found) == 1
    assert found[0]["lines"][0]["product"] == "Widget", "adopted the wrong sheet"
    assert "My Notes" in load_workbook(path).sheetnames


def test_an_unidentifiable_workbook_is_refused_rather_than_guessed(store):
    path = sales_path(store)
    wb = Workbook()
    wb.active.title = "One"
    wb.active.append(["nothing", "we", "recognise"])
    wb.create_sheet("Two").append(["also", "not", "ours"])
    wb.save(path)
    with pytest.raises(storage.DataIntegrityError) as err:
        store.bills(path)
    assert "Ledger" in str(err.value)


def test_a_legacy_ledger_is_upgraded_on_disk_at_startup(store):
    """Bill IDs must be frozen, not recomputed positionally on every read."""
    path = sales_path(store)
    wb = Workbook()
    ws = wb.active
    ws.append(storage._LEGACY_TXN_HEADERS)
    for product in ("Zeta", "Alpha"):
        ws.append(["01/01/2026", "B1", "P", "V", "A", product,
                   1, 10, 10, 20, 0, 13, 2.6, 22.6])
    ws.append(["02/01/2026", "B2", "P", "V", "A", "Solo",
               1, 50, 50, 50, 0, 13, 6.5, 56.5])
    wb.save(path)

    store.migrate_ledgers()
    saved = load_workbook(path)
    assert saved.sheetnames == [storage.SHEET_NAME], "not upgraded on disk"
    ws = saved[storage.SHEET_NAME]
    ids = [ws.cell(row=r, column=storage.C_BILLID + 1).value
           for r in range(2, ws.max_row + 1)]
    assert ids == [1, 1, 2], f"bill ids were not written: {ids}"

    before = [(b["bill_id"], b["totals"]["total"]) for b in store.bills(path)]
    # Now re-sort the sheet the way a user would; identity must not move.
    wb = load_workbook(path)
    ws = wb[storage.SHEET_NAME]
    data = [[ws.cell(row=r, column=c).value
             for c in range(1, len(storage.TXN_HEADERS) + 1)]
            for r in range(2, ws.max_row + 1)]
    data.sort(key=lambda row: str(row[storage.C_PRODUCT]))
    for r, row in enumerate(data, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    wb.save(path)
    assert [(b["bill_id"], b["totals"]["total"])
            for b in store.bills(path)] == before


def test_hand_added_rows_of_one_bill_are_grouped_not_counted_per_line(store):
    """Regression: every id-less row became its own bill, inflating totals."""
    path = sales_path(store)
    bill(store, bill_no="B1", qty=1, rate=100)
    wb = load_workbook(path)
    ws = wb[storage.SHEET_NAME]
    for product in ("Nut", "Bolt", "Washer"):        # one 3-line bill, no ids
        ws.append(["05/01/2026", "H1", "PAN9", "Hand", "Addr", product,
                   1, 10, 10, 30, 0, 13, 3.9, 33.9])
    wb.save(path)

    found = store.bills(path)
    assert len(found) == 2, f"expected 2 bills, got {len(found)}"
    hand = next(b for b in found if b["synthetic"])
    assert len(hand["lines"]) == 3, "a 3-line bill was split into 3 bills"
    assert hand["totals"]["total"] == 33.9
    assert round(sum(b["totals"]["total"] for b in found), 2) == 146.9

    store.rebuild_party_totals()
    rows = {r[storage.P_PAN]: r[storage.P_SALES]
            for r in store.read_rows(party_path(store), storage.PARTY_HEADERS)}
    assert rows["PAN9"] == 33.9, "the bill total was counted once per line"


def test_a_bill_cancelled_by_hand_entered_rows_is_not_voidable_again(store):
    """Regression: a synthetic canceller left voided_by falsy."""
    path = sales_path(store)
    target = bill(store, bill_no="B1", qty=1, rate=100)
    wb = load_workbook(path)
    ws = wb[storage.SHEET_NAME]
    ws.append(["01/01/2026", "VOID-B1", "AAAAA1111A", "Acme", "Addr", "Widget",
               -1, 100, -100, -100, 0, 13, -13, -113, None, "", target])
    wb.save(path)

    found = next(b for b in store.bills(path) if b["bill_id"] == target)
    assert found["voided_by"], "the cancellation was not detected"
    with pytest.raises(ValueError, match="already cancelled"):
        store.void_bill(True, target)


def test_rebuild_refuses_rather_than_erasing_a_user_column(store):
    bill(store, is_sale=False, product="Widget", qty=10, bill_no="P1")
    wb = load_workbook(stock_path(store))
    ws = wb[storage.SHEET_NAME]
    ws.cell(row=1, column=3, value="Reorder level")
    ws.cell(row=2, column=3, value=25)
    wb.save(stock_path(store))

    with pytest.raises(storage.DataIntegrityError) as err:
        store.rebuild_stock()
    assert "column" in str(err.value).lower()
    ws = load_workbook(stock_path(store))[storage.SHEET_NAME]
    assert ws.cell(row=2, column=3).value == 25, "the user's column was erased"


def test_void_reports_the_reversal_when_only_the_rebuild_fails(store,
                                                               monkeypatch):
    """A recorded reversal must never be reported as 'could not cancel'."""
    bill(store, is_sale=False, product="Widget", qty=50, bill_no="P1")
    sale_id = bill(store, is_sale=True, product="Widget", qty=10, bill_no="S1")

    def broken_rebuild():
        raise storage.StorageError("stock.xlsx is busy")

    monkeypatch.setattr(storage, "rebuild_stock", broken_rebuild)
    with pytest.raises(storage.RebuildFailedError) as err:
        store.void_bill(True, sale_id)
    assert "was cancelled" in str(err.value)
    assert err.value.bill_id is not None
    monkeypatch.undo()

    # the reversal really is in the ledger, so it must not be repeated
    with pytest.raises(ValueError, match="already cancelled"):
        store.void_bill(True, sale_id)


def test_a_second_instance_is_refused(store):
    store.acquire_single_instance_lock()
    try:
        code = (
            "import os, sys, storage\n"
            "try:\n"
            "    storage.acquire_single_instance_lock()\n"
            "except storage.AlreadyRunningError:\n"
            "    print('REFUSED'); sys.exit(0)\n"
            "print('ALLOWED'); sys.exit(1)\n")
        proc = subprocess.run(
            [sys.executable, "-c", code],
            cwd=os.path.dirname(os.path.abspath(storage.__file__)),
            env=dict(os.environ), capture_output=True, text=True)
        assert "REFUSED" in proc.stdout, (
            f"a second instance was allowed in: {proc.stdout}{proc.stderr}")
    finally:
        store.release_single_instance_lock()


def test_the_lock_is_released_so_the_next_run_can_start(store):
    store.acquire_single_instance_lock()
    store.release_single_instance_lock()
    store.acquire_single_instance_lock()      # must not raise
    store.release_single_instance_lock()


def test_header_text_is_shown_without_the_sanitizing_apostrophe(store):
    """The ' that guards against formula injection must not reach the user."""
    bill(store, bill_no="-77", vendor="-Traders", pan="", qty=1, rate=100)
    found = store.bills(sales_path(store))[0]
    assert found["header"]["bill"] == "-77"
    assert found["header"]["vendor"] == "-Traders"
