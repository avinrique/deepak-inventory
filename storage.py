"""
Storage layer for the Inventory Management app.

All Excel reading/writing and business rules live here so they can be tested
without a GUI. The UI (``inventory_app.py``) only calls these functions.

A single bill may contain many products. Each product becomes one row in the
sales/purchases workbook, all sharing the same Bill ID. Per-line figures
(Quantity, Rate, Amount = Quantity x Rate) differ per row; bill-level figures
(Subtotal, ECS, VAT %, VAT Amount, Total) are repeated on every row of the
bill so each row is fully self-contained and nothing reads as blank.

    sales.xlsx       - every sale (one row per product line)
    purchases.xlsx   - every purchase (one row per product line)
    stock.xlsx       - current quantity on hand, per product
    party.xlsx       - per-party totals (sales / purchases / combined)

The two ledgers are the source of truth. ``stock.xlsx`` and ``party.xlsx`` are
derived caches: they are maintained incrementally for speed, but can always be
recomputed from the ledgers with ``rebuild_stock`` / ``rebuild_party_totals``.
Nothing is ever deleted -- a mistaken bill is cancelled with ``void_bill``,
which appends a reversing entry and rebuilds the caches.

Durability
----------
Every workbook write goes to a temporary file first and is then swapped into
place with ``os.replace`` (atomic on Windows and POSIX), keeping the previous
version as ``<name>.bak``. A bill touches three workbooks, so the swaps are
recorded in a journal first: if the process dies midway, ``recover()`` finishes
the remaining swaps on the next launch. A bill is therefore either fully
applied or not applied at all. If recovery cannot finish, the journal is KEPT
and the caller is told -- the app never silently abandons a pending bill.

Numbers the app maintains are parsed strictly (``_stored_num``): a cell that
cannot be read as a number raises rather than defaulting to zero, because a
silent zero would overwrite a real balance.

Files live in a stable per-user folder (see ``data_dir``), deliberately OUTSIDE
the project/exe directory so they survive rebuilds, re-downloads and git
clean/checkpoint restores. Override with ``INVENTORY_DATA_DIR``.
"""

import os
import sys
import glob
import json
import math
import shutil
import logging
import platform
import tempfile
import subprocess
from datetime import datetime
from logging.handlers import RotatingFileHandler

from openpyxl import Workbook, load_workbook


# --------------------------------------------------------------------------- #
# Errors
# --------------------------------------------------------------------------- #
class StorageError(Exception):
    """Any storage failure that is not a file lock (disk full, bad path, ...)."""


class FileLockedError(StorageError):
    """A workbook could not be written because it is open in Excel/LibreOffice."""

    def __init__(self, path):
        self.path = path
        super().__init__(f"{os.path.basename(path)} is open in Excel/LibreOffice.")


class DataIntegrityError(StorageError):
    """A cell or sheet the app maintains holds something it cannot interpret.

    Raised instead of guessing, because guessing would overwrite real figures.
    """


class CommitInterruptedError(StorageError):
    """A commit was interrupted after at least one workbook had been swapped.

    The bill may already be in the ledger. The caller must NOT tell the user
    that nothing was saved.
    """


class RebuildFailedError(StorageError):
    """A bill was recorded, but recomputing the derived totals afterwards failed.

    Carries ``bill_id`` so the caller can say what DID happen instead of
    reporting the whole operation as a failure.
    """

    def __init__(self, message, bill_id=None):
        self.bill_id = bill_id
        super().__init__(message)


class AlreadyRunningError(StorageError):
    """Another copy of the app already has this data folder open."""


# --------------------------------------------------------------------------- #
# Schema
# --------------------------------------------------------------------------- #
SHEET_NAME = "Ledger"

TXN_HEADERS = [
    "Date", "Bill No", "PAN No", "Vendor Name", "Vendor Address",
    "Product Name", "Quantity", "Rate", "Amount",
    "Subtotal", "ECS", "VAT %", "VAT Amount", "Total",
    "Bill ID", "Entered At", "Voids Bill ID",
]
# Column indices into a row returned by read_rows() (0-based).
(C_DATE, C_BILL, C_PAN, C_VENDOR, C_ADDR,
 C_PRODUCT, C_QTY, C_RATE, C_AMOUNT,
 C_SUBTOTAL, C_ECS, C_VATPCT, C_VATAMT, C_TOTAL,
 C_BILLID, C_ENTERED, C_VOIDS) = range(len(TXN_HEADERS))

STOCK_HEADERS = ["Product Name", "Quantity"]
S_PRODUCT, S_QTY = 0, 1

PARTY_HEADERS = [
    "PAN No", "Vendor Name", "Vendor Address",
    "Total Sales", "Total Purchases", "Total Combined",
]
P_PAN, P_NAME, P_ADDR, P_SALES, P_PURCH, P_COMBINED = range(len(PARTY_HEADERS))

# Guard rail for a single money figure. Above this a float stops being exact to
# the paisa, and openpyxl writes non-finite values as blank cells.
MAX_MONEY = 1e12
QTY_DP = 4          # quantities are stored to 4 decimal places

_LEGACY_TXN_HEADERS = TXN_HEADERS[:14]
_NEW_TXN_COLUMNS = (15, 16, 17)     # 1-based: Bill ID, Entered At, Voids Bill ID
_TMP_GLOB = ".tmp-*.xlsx"

log = logging.getLogger("inventory")


# --------------------------------------------------------------------------- #
# Storage location
# --------------------------------------------------------------------------- #
def _user_data_base() -> str:
    """Per-user, OS-appropriate folder for application data.

    Deliberately OUTSIDE the project/exe folder so the data can never be wiped
    by a git operation (the project ``.gitignore`` excludes ``inventory_data``),
    a rebuild that recreates ``dist``, a re-download, or a checkpoint/clean
    that restores the working tree to a snapshot.

    Override with ``INVENTORY_DATA_DIR``. That value is treated as a BASE
    folder: the workbooks land in ``<INVENTORY_DATA_DIR>/inventory_data``.
    ``~`` and ``%VARS%`` are expanded.
    """
    override = os.environ.get("INVENTORY_DATA_DIR")
    if override:
        return os.path.expanduser(os.path.expandvars(override))

    system = platform.system()
    home = os.path.expanduser("~")
    if system == "Windows":
        root = os.environ.get("APPDATA") or os.path.join(home, "AppData", "Roaming")
    elif system == "Darwin":
        root = os.path.join(home, "Library", "Application Support")
    else:
        root = os.environ.get("XDG_DATA_HOME") or os.path.join(home, ".local", "share")
    return os.path.join(root, "InventoryManagement")


def _folder() -> str:
    """The data folder, created if needed. No migration, no side effects.

    Kept separate from ``data_dir`` so ``setup_logging`` can attach the log
    handler BEFORE the legacy migration runs -- otherwise the migration's own
    log lines are written to a handler that does not exist yet.
    """
    folder = os.path.join(_user_data_base(), "inventory_data")
    os.makedirs(folder, exist_ok=True)
    return folder


def _legacy_data_dir() -> str:
    """The old location: ``inventory_data`` next to the exe (frozen) or module."""
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "inventory_data")


_data_dir_cache = None
_legacy_copied = False


def data_dir() -> str:
    """Folder where the Excel files live (stable, per-user, outside the repo).

    Resolved lazily and cached, so importing this module never touches the
    filesystem -- an unwritable folder must surface as a handled error inside
    the running app, not as an import-time crash with no window to show it in.

    One-time migration: if the new folder has no workbooks yet but the old
    ``inventory_data`` folder next to the app does, copy them across so an
    existing user keeps their data. The copy is staged and renamed, so an
    interrupted migration leaves no half-written workbook behind.
    """
    global _data_dir_cache, _legacy_copied
    if _data_dir_cache is not None:
        return _data_dir_cache

    folder = _folder()
    legacy = _legacy_data_dir()
    if os.path.abspath(legacy) != os.path.abspath(folder) and os.path.isdir(legacy):
        for name in ("sales.xlsx", "purchases.xlsx", "stock.xlsx", "party.xlsx"):
            src, dst = os.path.join(legacy, name), os.path.join(folder, name)
            if os.path.exists(src) and not os.path.exists(dst):
                staged = dst + ".tmp-migrate"
                try:
                    shutil.copy2(src, staged)
                    os.replace(staged, dst)
                except OSError as exc:
                    log.warning("legacy migration of %s failed: %s", name, exc)
                    try:
                        os.remove(staged)
                    except OSError:
                        pass
                    continue
                _legacy_copied = True
                log.info("migrated legacy workbook %s", name)

    _data_dir_cache = folder
    return folder


def reset_data_dir() -> None:
    """Forget the resolved data folder so ``INVENTORY_DATA_DIR`` is re-read.

    Exists for tests, which must be able to repoint storage at a scratch
    folder. Nothing in the app calls it.
    """
    global _data_dir_cache, _legacy_copied
    _data_dir_cache = None
    _legacy_copied = False


def legacy_data_was_migrated() -> bool:
    """True if this run copied workbooks in from the old location.

    The app rebuilds the derived caches once when this is true: files written
    by the previous version may carry duplicate product/party rows left by a
    bug it had, and the rebuild is what heals them.
    """
    return _legacy_copied


def __getattr__(name: str) -> str:
    """Resolve the file-path constants lazily (PEP 562).

    ``db.SALES_FILE`` still reads like a constant at the call site but does not
    force the data folder to exist at import time.
    """
    paths = {
        "DATA_DIR": "",
        "SALES_FILE": "sales.xlsx",
        "PURCHASES_FILE": "purchases.xlsx",
        "STOCK_FILE": "stock.xlsx",
        "PARTY_FILE": "party.xlsx",
    }
    if name in paths:
        return os.path.join(data_dir(), paths[name]) if paths[name] else data_dir()
    raise AttributeError(f"module {__name__!r} has no attribute {name!r}")


def ledger_files() -> tuple:
    """(sales, purchases) paths -- the two source-of-truth workbooks."""
    return (os.path.join(data_dir(), "sales.xlsx"),
            os.path.join(data_dir(), "purchases.xlsx"))


# --------------------------------------------------------------------------- #
# Logging
# --------------------------------------------------------------------------- #
def setup_logging(level=logging.INFO) -> str:
    """Send the app log to ``<data folder>/app.log``. Returns the log path.

    Called first at startup, before ``data_dir()``, so that the legacy
    migration's log lines are actually captured. A packaged ``--windowed``
    build has no console, so this file is the only record of what went wrong
    in the field.
    """
    path = os.path.join(_folder(), "app.log")
    if not any(isinstance(h, RotatingFileHandler) for h in log.handlers):
        handler = RotatingFileHandler(path, maxBytes=512_000, backupCount=3,
                                      encoding="utf-8")
        handler.setFormatter(logging.Formatter(
            "%(asctime)s %(levelname)-7s %(message)s"))
        log.addHandler(handler)
        log.setLevel(level)
        log.propagate = False
    return path


# --------------------------------------------------------------------------- #
# Value normalisation
# --------------------------------------------------------------------------- #
def _safe_text(value) -> str:
    """Neutralize spreadsheet formula injection in a free-text cell.

    A value beginning with = + - @ is prefixed with ' so Excel/LibreOffice
    stores it as literal text rather than evaluating it as a formula. openpyxl
    really does turn a leading "=" string into a live formula cell, so this
    guard matters.

    The apostrophe becomes part of the stored string, so every lookup MUST
    compare with ``_key`` rather than with the raw text -- see ``_key``.
    """
    s = "" if value is None else str(value)
    return "'" + s if s[:1] in ("=", "+", "-", "@") else s


def _key(value) -> str:
    """Canonical match key for a name/PAN/bill-number cell.

    Strips the apostrophe ``_safe_text`` may have added, then trims and
    lowercases, so the sanitized stored form and the raw typed form compare
    equal. Every read/write pair in this module keys off this function; keying
    off raw text on one side and sanitized text on the other silently forks a
    product or party into a new row on every transaction.
    """
    s = "" if value is None else str(value)
    if s[:2] in ("'=", "'+", "'-", "'@"):
        s = s[1:]
    return s.strip().lower()


def _display(value) -> str:
    """The text the user originally typed, with any sanitizing apostrophe gone."""
    s = "" if value is None else str(value).strip()
    return s[1:] if s[:2] in ("'=", "'+", "'-", "'@") else s


def num(value) -> float:
    """Parse USER INPUT into a float; blanks/junk/non-finite become 0.

    A trailing percent sign and thousands commas are tolerated, so "13%" and
    "1,030" parse as 13 and 1030. Deliberately permissive -- this is for form
    fields. For any number the app itself maintains use ``_stored_num``, which
    refuses rather than defaulting to zero.
    """
    if value is None:
        return 0.0
    try:
        result = float(str(value).replace(",", "").replace("%", "").strip() or 0)
    except ValueError:
        return 0.0
    return result if math.isfinite(result) else 0.0


def _stored_num(value, path: str, row: int, column: str) -> float:
    """Parse a number the app itself maintains. Blank is 0; junk is an error.

    Silently reading an unparseable balance as 0 would overwrite it with a
    wrong absolute figure on the next write, so this raises instead. Used for
    every app-maintained numeric column in BOTH the ledgers and the caches.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return 0.0
    if isinstance(value, bool):
        raise DataIntegrityError(
            f"{os.path.basename(path)} row {row}, {column} holds TRUE/FALSE "
            f"instead of a number.")
    if isinstance(value, (int, float)):
        if math.isfinite(value):
            return float(value)
        raise DataIntegrityError(
            f"{os.path.basename(path)} row {row}, {column} is not a finite "
            f"number. Fix that cell in Excel, then try again.")
    text = str(value).strip()
    try:
        result = float(text.replace(",", ""))
    except ValueError:
        hint = (" If that is a formula, replace it with a plain number."
                if text.startswith("=") else "")
        raise DataIntegrityError(
            f"{os.path.basename(path)} row {row}, {column} contains {text!r}, "
            f"which is not a number.{hint} Fix that cell in Excel, then try "
            f"again.") from None
    if not math.isfinite(result):
        raise DataIntegrityError(
            f"{os.path.basename(path)} row {row}, {column} is not a finite "
            f"number.")
    return result


def _money(value: float, what: str) -> float:
    """Round to paisa and reject values that cannot be stored faithfully."""
    if not math.isfinite(value):
        raise ValueError(f"{what} is not a finite number.")
    if abs(value) > MAX_MONEY:
        raise ValueError(
            f"{what} is too large to record (limit {MAX_MONEY:,.0f}). "
            f"Check the quantity and rate.")
    return round(value, 2)


# --------------------------------------------------------------------------- #
# Business rules (kept out of the GUI so they can be tested headlessly)
# --------------------------------------------------------------------------- #
def line_amount(qty: float, rate: float) -> float:
    """Amount for one product line. Raises ValueError if it cannot be stored."""
    return _money(qty * rate, "Amount")


def bill_totals(lines: list, ecs: float, vat_pct: float) -> dict:
    """Bill-level figures from the product lines.

    Subtotal = sum of line amounts, VAT Amount = VAT % of Subtotal,
    Total = Subtotal + ECS + VAT Amount. The single definition of this
    arithmetic -- the GUI displays what this returns rather than recomputing.
    """
    subtotal = _money(sum(ln["amount"] for ln in lines), "Subtotal")
    vat_amount = _money(subtotal * vat_pct / 100.0, "VAT Amount")
    total = _money(subtotal + ecs + vat_amount, "Total")
    return {"subtotal": subtotal, "ecs": _money(ecs, "ECS"),
            "vat_pct": vat_pct, "vat_amount": vat_amount, "total": total}


def shortages(lines: list) -> list:
    """Products whose TOTAL quantity across the bill exceeds stock on hand.

    Quantities are summed per product first: two lines of 5 against 8 on hand
    is a shortage even though neither line alone is. The sum is rounded to the
    precision stock is stored at, so selling back exactly what was bought does
    not trip a phantom warning on a float tail.
    Returns [{"product", "have", "selling"}] for the products that fall short.
    """
    wanted = {}
    for ln in lines:
        k = _key(ln["product"])
        entry = wanted.setdefault(k, {"product": ln["product"], "selling": 0.0})
        entry["selling"] += ln["qty"]
    out = []
    for entry in wanted.values():
        selling = round(entry["selling"], QTY_DP)
        have = stock_on_hand(entry["product"])
        if selling > have:
            out.append({"product": entry["product"], "have": have,
                        "selling": selling})
    return out


# --------------------------------------------------------------------------- #
# Workbook plumbing
# --------------------------------------------------------------------------- #
def _row1(ws, width: int) -> list:
    return [ws.cell(row=1, column=c).value for c in range(1, width + 1)]


def _is_ledger_sheet(ws, headers: list) -> bool:
    """Does this sheet's header row look like the data we manage?"""
    if headers is TXN_HEADERS:
        return (_row1(ws, len(TXN_HEADERS)) == TXN_HEADERS
                or _row1(ws, len(_LEGACY_TXN_HEADERS)) == _LEGACY_TXN_HEADERS)
    return _row1(ws, len(headers)) == headers


def _sheet(wb, headers: list = None):
    """The worksheet holding our data, identified by NAME then by CONTENT.

    Never ``wb.active``: if the user adds a tab in Excel and leaves it
    selected, Excel stores that as the active sheet and the app would silently
    read and write the wrong one, orphaning the real ledger.

    Nor blindly ``worksheets[0]``: on a pre-upgrade workbook the ledger is on a
    sheet called "Sheet", and if the user has since added their own tab and
    dragged it to the front, renaming position 0 to "Ledger" would adopt THEIR
    tab and orphan the real data.
    """
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    if headers is not None:
        for ws in wb.worksheets:
            if _is_ledger_sheet(ws, headers):
                ws.title = SHEET_NAME
                return ws
    if len(wb.worksheets) == 1:
        ws = wb.worksheets[0]
        ws.title = SHEET_NAME
        return ws
    raise DataIntegrityError(
        f"None of the sheets ({', '.join(wb.sheetnames)}) has the expected "
        f"header row, so the app cannot tell which one holds its data. Rename "
        f"the correct sheet to '{SHEET_NAME}' in Excel, then try again.")


def _migrate(ws, headers: list, path: str) -> bool:
    """Bring an older workbook up to the current schema. True if changed."""
    changed = False
    if ws.title != SHEET_NAME:
        ws.title = SHEET_NAME
        changed = True

    if headers is not TXN_HEADERS:
        # stock/party schemas have never changed; only the header row can be
        # missing, and only on an otherwise empty sheet may we write it.
        if _row1(ws, len(headers)) != headers:
            if ws.max_row > 1:
                raise DataIntegrityError(
                    f"{os.path.basename(path)} does not have the expected "
                    f"header row ({', '.join(headers)}). Restore it in Excel, "
                    f"or delete the file and press Recalculate from Ledgers.")
            for c, name in enumerate(headers, start=1):
                ws.cell(row=1, column=c, value=name)
            changed = True
        return changed

    if _row1(ws, len(TXN_HEADERS)) == TXN_HEADERS:
        return changed

    # Only treat it as the old 14-column layout if the three columns we are
    # about to claim are genuinely unused -- otherwise we would overwrite the
    # owner's own column O and everything in it.
    legacy_layout = _row1(ws, 14) == _LEGACY_TXN_HEADERS
    columns_free = not any(
        ws.cell(row=r, column=c).value not in (None, "")
        for r in range(1, ws.max_row + 1) for c in _NEW_TXN_COLUMNS)
    if not (legacy_layout and columns_free):
        raise DataIntegrityError(
            f"{os.path.basename(path)} does not have a header row this version "
            f"recognises, or already uses columns O-Q for something else. The "
            f"app will not touch it. Expected row 1 to read: "
            f"{', '.join(TXN_HEADERS)}.")

    log.info("upgrading %s: adding Bill ID / Entered At / Voids Bill ID",
             os.path.basename(path))
    for c, name in enumerate(TXN_HEADERS, start=1):
        ws.cell(row=1, column=c, value=name)

    # Backfill Bill ID for historical rows. The old format had no bill
    # identity, so reconstruct it the only way the data allows: rows sharing
    # the same header details AND the same bill-level money belonged to one
    # bill. Keyed through a dict rather than by comparing with the previous
    # row, so a sheet the user has sorted in Excel still groups correctly.
    # Two same-day blank-Bill-No bills with identical totals are genuinely
    # indistinguishable in the old format; that case is logged, not hidden.
    groups, assigned = {}, {}
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for r in range(2, ws.max_row + 1):
        base = tuple(_key(ws.cell(row=r, column=c).value)
                     for c in list(range(1, 6)) + list(range(10, 15)))
        if not any(base):
            continue
        bill_id = groups.setdefault(base, len(groups) + 1)
        assigned.setdefault(bill_id, 0)
        assigned[bill_id] += 1
        ws.cell(row=r, column=C_BILLID + 1, value=bill_id)
        ws.cell(row=r, column=C_ENTERED + 1, value=stamp)
    log.info("backfilled %d historical bill id(s) across %d row(s)",
             len(groups), sum(assigned.values()))
    return True


def _blank_workbook(headers: list):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(headers)
    return wb


def _ensure_file(path: str, headers: list) -> None:
    """Create the workbook with a header row if it does not exist yet."""
    if os.path.exists(path):
        return
    _save(_blank_workbook(headers), path)


def _open(path: str, headers: list, persist_migration: bool):
    """(workbook, worksheet), creating and migrating the file as needed.

    Always opened with ``data_only=False``. Loading cached values instead would
    turn a formula the user typed into an app-managed column into ``None``,
    which reads as a silent zero and then overwrites the real balance. Keeping
    the formula text means ``_stored_num`` can refuse it by name, and it also
    means a save never rewrites the user's formulas as flat values.
    """
    _ensure_file(path, headers)
    try:
        wb = load_workbook(path, data_only=False)
    except PermissionError as exc:
        raise FileLockedError(path) from exc
    except OSError as exc:
        raise StorageError(
            f"Could not read {os.path.basename(path)}: {exc}") from exc
    except Exception as exc:  # noqa: BLE001 - a corrupt zip lands here
        backup = path + ".bak"
        hint = (f" A previous copy is available as {os.path.basename(backup)}."
                if os.path.exists(backup) else "")
        raise StorageError(
            f"{os.path.basename(path)} could not be opened -- the file may be "
            f"damaged.{hint} ({exc})") from exc
    ws = _sheet(wb, headers)
    if _migrate(ws, headers, path) and persist_migration:
        _save(wb, path)
    return wb, ws


def _load_ro(path: str, headers: list):
    """Open for reading. A read never rewrites the user's file."""
    return _open(path, headers, persist_migration=False)


def _load_rw(path: str, headers: list):
    """Open for read-modify-write, persisting any schema upgrade."""
    return _open(path, headers, persist_migration=True)


# --------------------------------------------------------------------------- #
# Durable writes
# --------------------------------------------------------------------------- #
def _journal_path() -> str:
    return os.path.join(data_dir(), ".commit-journal.json")


def _lock_path() -> str:
    return os.path.join(data_dir(), ".instance.lock")


_lock_handle = None


def acquire_single_instance_lock() -> None:
    """Take an advisory lock on the data folder for this process's lifetime.

    These are plain files with no record locking, so two copies of the app
    pointed at one folder would each read-modify-write the same workbooks and
    silently lose the other's bill. One instance also has to be able to sweep
    stale staging files without deleting another's in-flight ones.

    Raises AlreadyRunningError if another instance holds it.
    """
    global _lock_handle
    if _lock_handle is not None:
        return
    path = _lock_path()
    handle = open(path, "a+b")
    try:
        if os.name == "nt":
            import msvcrt
            handle.seek(0)
            msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
        else:
            import fcntl
            fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
    except OSError as exc:
        handle.close()
        raise AlreadyRunningError(
            "Another copy of Inventory Management is already using this data "
            "folder. Close it first — running two copies at once would lose "
            "bills.") from exc
    except ImportError:                 # no locking available: carry on
        log.warning("file locking unavailable on this platform")
        _lock_handle = handle
        return
    _lock_handle = handle


def release_single_instance_lock() -> None:
    """Drop the advisory lock. Safe to call when it was never taken."""
    global _lock_handle
    if _lock_handle is None:
        return
    try:
        _lock_handle.close()
    except OSError:
        pass
    _lock_handle = None


def _sync_dir(folder: str) -> None:
    """Best-effort flush of a directory entry. No-op where unsupported."""
    if not hasattr(os, "O_DIRECTORY"):
        return                                   # Windows: not available
    try:
        fd = os.open(folder, os.O_DIRECTORY)
        try:
            os.fsync(fd)
        finally:
            os.close(fd)
    except OSError:
        pass


def _stage(wb, dest: str) -> str:
    """Write ``wb`` to a temp file beside ``dest``; return the temp path."""
    folder = os.path.dirname(dest) or "."
    fd, tmp = tempfile.mkstemp(prefix=".tmp-", suffix=".xlsx", dir=folder)
    os.close(fd)
    try:
        wb.save(tmp)
        with open(tmp, "rb+") as fh:
            fh.flush()
            os.fsync(fh.fileno())
    except PermissionError as exc:
        _discard(tmp)
        raise FileLockedError(dest) from exc
    except OSError as exc:
        _discard(tmp)
        raise StorageError(
            f"Could not write {os.path.basename(dest)}: {exc}") from exc
    except Exception:
        _discard(tmp)
        raise
    return tmp


def _discard(tmp: str) -> bool:
    """Remove a file, reporting (and logging) failure rather than hiding it."""
    try:
        os.remove(tmp)
        return True
    except FileNotFoundError:
        return True
    except OSError as exc:
        log.warning("could not remove %s: %s", tmp, exc)
        return False


def _write_journal(pairs: list, keep_backup: bool) -> None:
    """Write the commit journal atomically. Raises OSError on failure.

    Atomic because this file is the ONLY record of an unfinished commit: a
    half-written journal reads as damaged, and the recovery path for a damaged
    journal throws the staged workbooks away.
    """
    journal = _journal_path()
    folder = os.path.dirname(journal) or "."
    fd, tmp = tempfile.mkstemp(prefix=".tmp-journal-", suffix=".json", dir=folder)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            json.dump({"pairs": pairs, "keep_backup": keep_backup,
                       "written": datetime.now().isoformat(timespec="seconds")},
                      fh)
            fh.flush()
            os.fsync(fh.fileno())
        os.replace(tmp, journal)
        _sync_dir(folder)
    except OSError:
        _discard(tmp)
        raise


def _write_backup(dest: str) -> None:
    """Refresh ``dest.bak`` atomically, so it is never observed half-written."""
    if not os.path.exists(dest):
        return
    folder = os.path.dirname(dest) or "."
    fd, tmp = tempfile.mkstemp(prefix=".tmp-", suffix=".xlsx", dir=folder)
    os.close(fd)
    try:
        shutil.copy2(dest, tmp)
        os.replace(tmp, dest + ".bak")
    except OSError as exc:      # a missing backup must not block the write
        _discard(tmp)
        log.warning("could not refresh %s.bak: %s", os.path.basename(dest), exc)


def _swap(tmp: str, dest: str, keep_backup: bool = True) -> None:
    """Put ``tmp`` in place of ``dest``, optionally keeping the old ``.bak``."""
    if keep_backup:
        _write_backup(dest)
    os.replace(tmp, dest)


def _commit_staged(pairs: list, keep_backup: bool = True) -> None:
    """Swap several staged files into place, recording intent first.

    The individual ``os.replace`` calls are atomic, but a set of them is not:
    a crash between two swaps would leave the ledger updated and the caches
    stale. The journal makes that state recoverable -- ``recover()`` finishes
    the outstanding swaps on the next launch.
    """
    journal = _journal_path()
    if os.path.exists(journal):
        # Starting a new commit here would truncate the record of the previous
        # unfinished one, stranding its staged files forever.
        for tmp, _ in pairs:
            _discard(tmp)
        raise StorageError(
            "A previous save was interrupted and has not been completed yet. "
            "Close and reopen the app -- it finishes the pending save on "
            "startup -- then try again.")
    try:
        _write_journal(pairs, keep_backup)
    except OSError as exc:
        for tmp, _ in pairs:
            _discard(tmp)
        raise StorageError(f"Could not start the save: {exc}") from exc

    done, remaining = 0, list(pairs)
    try:
        for tmp, dest in pairs:
            _swap(tmp, dest, keep_backup)
            done += 1
            # Narrow the journal as we go, so a staged file that is listed but
            # missing is unambiguously LOST rather than "already swapped".
            remaining = remaining[1:]
            _write_journal(remaining, keep_backup)
    except OSError as exc:
        if done:
            raise CommitInterruptedError(
                f"The save was interrupted after {done} of {len(pairs)} files "
                f"were written ({exc}). It will be completed the next time the "
                f"app starts.") from exc
        _discard(journal)
        for tmp, _ in pairs:
            _discard(tmp)
        raise StorageError(f"The save could not be started: {exc}") from exc
    if not _discard(journal):
        raise StorageError(
            f"The save completed, but the record of it could not be removed "
            f"({journal}). Delete that file, then restart the app.")
    _sync_dir(os.path.dirname(journal) or ".")


def _sweep_orphans(keep: set = frozenset()) -> int:
    """Delete staging files no live journal refers to. Returns how many."""
    removed = 0
    for tmp in glob.glob(os.path.join(data_dir(), _TMP_GLOB)):
        if tmp in keep:
            continue
        _discard(tmp)
        removed += 1
    if removed:
        log.info("removed %d orphaned staging file(s)", removed)
    return removed


def recover() -> dict:
    """Finish a commit that was interrupted by a crash.

    Returns {"recovered": bool, "failed": [paths], "message": str}. If any swap
    cannot be completed the journal is KEPT, the rebuilds are skipped, and the
    caller is expected to show ``message`` to the user -- abandoning a pending
    bill silently is exactly the failure this whole mechanism exists to avoid.

    Called once at startup, before anything reads the workbooks.
    """
    journal = _journal_path()
    if not os.path.exists(journal):
        _sweep_orphans()
        return {"recovered": False, "failed": [], "message": ""}

    try:
        with open(journal, encoding="utf-8") as fh:
            data = json.load(fh)
        pairs = data.get("pairs", [])
        keep_backup = data.get("keep_backup", True)
    except (OSError, ValueError) as exc:
        log.error("unreadable commit journal, removing it: %s", exc)
        _discard(journal)
        _sweep_orphans()
        return {"recovered": False, "failed": [],
                "message": "A record of an interrupted save was damaged and "
                           "could not be replayed. Press Recalculate from "
                           "Ledgers on the Stock and Parties pages."}

    log.warning("recovering an interrupted save (%d file(s))", len(pairs))
    done, failed, lost, pending = 0, [], [], []
    for tmp, dest in pairs:
        if not os.path.exists(tmp):
            # The journal is narrowed after every successful swap, so a pair
            # still listed whose staged file is gone was LOST, not applied.
            log.error("staged file for %s is missing; that change is lost", dest)
            lost.append(dest)
            continue
        try:
            _swap(tmp, dest, keep_backup)
            done += 1
        except OSError as exc:
            log.error("could not finish swapping %s: %s", dest, exc)
            failed.append(dest)
            pending.append([tmp, dest])

    if failed:
        # Keep the journal, narrowed to what is still outstanding, so the next
        # launch can try again once the user closes the file in Excel.
        try:
            _write_journal(pending, keep_backup)
        except OSError as exc:
            # Do NOT leave a truncated journal: the next launch would call it
            # damaged and sweep away the staged files it still points at.
            log.error("could not narrow the commit journal, leaving it as is: %s",
                      exc)
        names = ", ".join(os.path.basename(p) for p in failed)
        message = (f"A previous save could not be finished because {names} "
                   f"could not be written. Close it in Excel/LibreOffice and "
                   f"restart the app.")
        if done:
            message += (" Part of that save did land, so the Stock and Parties "
                        "pages are out of date until it completes.")
        return {"recovered": done > 0, "failed": failed, "message": message}

    if not _discard(journal):
        return {"recovered": done > 0, "failed": [],
                "message": f"An interrupted save was completed, but its record "
                           f"could not be removed. Until you delete this file "
                           f"the app cannot save anything:\n\n{journal}"}
    _sweep_orphans()
    log.warning("recovery completed %d swap(s); rebuilding derived totals", done)

    message = ""
    if lost:
        names = ", ".join(os.path.basename(p) for p in lost)
        message = (f"An interrupted save could NOT be completed — the pending "
                   f"changes to {names} were lost. Check the Sales and "
                   f"Purchases pages and re-enter the last bill if it is "
                   f"missing.")
    try:
        rebuild_stock()
        rebuild_party_totals()
    except StorageError as exc:
        log.error("post-recovery rebuild failed: %s", exc)
        message = (f"{message} " if message else "") + (
            f"Stock and party totals could not be recalculated: {exc} Press "
            f"Recalculate from Ledgers on the Stock and Parties pages.")
    return {"recovered": done > 0 and not lost, "failed": [],
            "message": message}


def _save(wb, path: str, keep_backup: bool = True) -> None:
    """Save one workbook atomically, keeping the previous copy as ``.bak``."""
    _commit_staged([(_stage(wb, path), path)], keep_backup)


def assert_writable(*paths) -> None:
    """Best-effort pre-check that the target workbooks are not locked.

    A courtesy so the common case -- the owner left a workbook open in Excel --
    fails immediately with a clear message. It is NOT what makes a commit safe:
    the file could be locked a millisecond later, and on POSIX this probe
    cannot see another process's handle at all. Atomicity comes from the
    staged-write journal in ``_commit_staged``.
    """
    for p in paths:
        if os.path.exists(p):
            try:
                with open(p, "r+b"):
                    pass
            except PermissionError as exc:
                raise FileLockedError(p) from exc
            except OSError as exc:
                raise StorageError(
                    f"Could not open {os.path.basename(p)}: {exc}") from exc


def backup_daily() -> str:
    """Copy the four workbooks into ``backups/<YYYY-MM-DD>`` once per day.

    Cheap insurance: the app rewrites these files constantly, and a mistake
    noticed on Thursday is unrecoverable if Monday's version is gone.
    """
    folder = os.path.join(data_dir(), "backups",
                          datetime.now().strftime("%Y-%m-%d"))
    if os.path.isdir(folder):
        return folder
    os.makedirs(folder, exist_ok=True)
    for name in ("sales.xlsx", "purchases.xlsx", "stock.xlsx", "party.xlsx"):
        src = os.path.join(data_dir(), name)
        if os.path.exists(src):
            try:
                shutil.copy2(src, os.path.join(folder, name))
            except OSError as exc:
                log.warning("daily backup of %s failed: %s", name, exc)
    log.info("daily backup written to %s", folder)
    return folder


# --------------------------------------------------------------------------- #
# Reads
# --------------------------------------------------------------------------- #
def _iter_rows(path: str, headers: list):
    """Yield (worksheet_row_number, values) for every non-blank data row."""
    wb, ws = _load_ro(path, headers)
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if any(v is not None and str(v).strip() != "" for v in row):
            yield r, row


def read_rows(path: str, headers: list) -> list:
    """Return all data rows (excluding the header) as lists."""
    return [row for _, row in _iter_rows(path, headers)]


def stock_on_hand(product: str) -> float:
    """Current quantity for a product (0 if unknown).

    Sums every row matching the product, so duplicate rows left behind by an
    older version's matching bug are read correctly rather than half-ignored.
    """
    path = os.path.join(data_dir(), "stock.xlsx")
    key = _key(product)
    total, found = 0.0, False
    for r, row in _iter_rows(path, STOCK_HEADERS):
        if _key(row[S_PRODUCT]) == key:
            total += _stored_num(row[S_QTY], path, r, "Quantity")
            found = True
    return round(total, QTY_DP) if found else 0.0


def product_names() -> list:
    """Distinct product names currently known in stock (for the dropdown)."""
    path = os.path.join(data_dir(), "stock.xlsx")
    names, seen = [], set()
    for _, row in _iter_rows(path, STOCK_HEADERS):
        text, key = _display(row[S_PRODUCT]), _key(row[S_PRODUCT])
        if text and key not in seen:
            seen.add(key)
            names.append(text)
    return sorted(names, key=str.lower)


def bill_exists(path: str, bill_no: str) -> bool:
    """True if a non-blank Bill No already appears in this ledger."""
    key = _key(bill_no)
    if not key:
        return False
    return any(_key(row[C_BILL]) == key for _, row in _iter_rows(path, TXN_HEADERS))


def next_bill_id(ws) -> int:
    """One past the highest Bill ID present in this worksheet."""
    highest = 0
    for r in range(2, ws.max_row + 1):
        value = ws.cell(row=r, column=C_BILLID + 1).value
        try:
            highest = max(highest, int(float(value)))
        except (TypeError, ValueError):
            continue
    return highest + 1


def bills(path: str) -> list:
    """Every bill in a ledger, grouped by Bill ID rather than by row order.

    Returns [{"bill_id", "header", "lines", "totals", "voids", "voided_by",
    "synthetic"}]. Grouping on a stored identity means the result is correct no
    matter how the user has sorted the sheet in Excel, and two same-day bills
    for one party with a blank Bill No stay separate.

    A row whose Bill ID is missing or unreadable (hand-added in Excel) is
    grouped with the other id-less rows that share its header details AND its
    bill-level money -- the same reconstruction ``_migrate`` uses. Such a group
    is flagged ``synthetic``: merging them all under one id would drop money
    from the grand total, and treating each row as its own bill would count a
    multi-line bill's total once per line.
    """
    grouped, order = {}, []
    for r, row in _iter_rows(path, TXN_HEADERS):
        try:
            gid, synthetic = ("id", int(float(row[C_BILLID]))), False
        except (TypeError, ValueError):
            gid = ("sig", tuple(_key(row[c]) for c in (
                C_DATE, C_BILL, C_PAN, C_VENDOR, C_ADDR,
                C_SUBTOTAL, C_ECS, C_VATPCT, C_VATAMT, C_TOTAL)))
            synthetic = True
        if gid not in grouped:
            grouped[gid] = {
                "bill_id": None if synthetic else gid[1],
                "row": r,
                "synthetic": synthetic,
                "header": {"date": _display(row[C_DATE]),
                           "bill": _display(row[C_BILL]),
                           "pan": _display(row[C_PAN]),
                           "vendor": _display(row[C_VENDOR]),
                           "address": _display(row[C_ADDR]),
                           "entered": row[C_ENTERED] or ""},
                "lines": [],
                "totals": {
                    "subtotal": _stored_num(row[C_SUBTOTAL], path, r, "Subtotal"),
                    "ecs": _stored_num(row[C_ECS], path, r, "ECS"),
                    "vat_pct": _stored_num(row[C_VATPCT], path, r, "VAT %"),
                    "vat_amount": _stored_num(row[C_VATAMT], path, r, "VAT Amount"),
                    "total": _stored_num(row[C_TOTAL], path, r, "Total")},
                "voids": row[C_VOIDS],
                "voided_by": None,
            }
            order.append(gid)
        grouped[gid]["lines"].append({
            "product": _display(row[C_PRODUCT]),
            "qty": _stored_num(row[C_QTY], path, r, "Quantity"),
            "rate": _stored_num(row[C_RATE], path, r, "Rate"),
            "amount": _stored_num(row[C_AMOUNT], path, r, "Amount")})

    by_id = {b["bill_id"]: b for b in grouped.values() if b["bill_id"] is not None}
    for bill in grouped.values():
        try:
            target = int(float(bill["voids"]))
        except (TypeError, ValueError):
            continue
        if target in by_id:
            # Never store a falsy marker: a synthetic canceller with bill_id
            # None would leave voided_by falsy and let the bill be voided again.
            by_id[target]["voided_by"] = (
                bill["bill_id"] if bill["bill_id"] is not None
                else f"the hand-entered rows at row {bill['row']}")

    return [grouped[g] for g in sorted(
        order, key=lambda g: (grouped[g]["bill_id"] is None, grouped[g]["bill_id"]
                              if grouped[g]["bill_id"] is not None
                              else grouped[g]["row"]))]


# --------------------------------------------------------------------------- #
# Writes -- one bill, applied atomically
# --------------------------------------------------------------------------- #
def _apply_stock(ws, path: str, product: str, qty: float, add: bool) -> float:
    """Add to (purchase) or subtract from (sale) stock, in memory.

    Folds duplicate rows for one product: the first row absorbs the total and
    any others are zeroed, healing files damaged by an older version.
    """
    key = _key(product)
    matches = [r for r in range(2, ws.max_row + 1)
               if _key(ws.cell(row=r, column=S_PRODUCT + 1).value) == key]
    if matches:
        current = sum(_stored_num(ws.cell(row=r, column=S_QTY + 1).value,
                                  path, r, "Quantity") for r in matches)
        new_qty = round(current + qty if add else current - qty, QTY_DP)
        ws.cell(row=matches[0], column=S_QTY + 1, value=new_qty)
        for r in matches[1:]:
            ws.cell(row=r, column=S_QTY + 1, value=0)
        return new_qty
    new_qty = round(qty if add else -qty, QTY_DP)
    ws.append([_safe_text(str(product).strip()), new_qty])
    return new_qty


def _apply_party(ws, path: str, pan: str, name: str, address: str,
                 total: float, is_sale: bool) -> None:
    """Update a party's running totals, in memory.

    Keyed by PAN when present; otherwise by Vendor Name + Address together
    (so two different blank-PAN vendors are not merged). A bill with neither
    PAN nor name is bucketed under a single "Unknown" party.
    """
    pan = str(pan or "")
    name = str(name or "")
    address = str(address or "")
    key_pan, key_name, key_addr = _key(pan), _key(name), _key(address)
    if not key_pan and not key_name:
        name, key_name = "Unknown", "unknown"

    for r in range(2, ws.max_row + 1):
        rp = _key(ws.cell(row=r, column=P_PAN + 1).value)
        rn = _key(ws.cell(row=r, column=P_NAME + 1).value)
        ra = _key(ws.cell(row=r, column=P_ADDR + 1).value)
        hit = (rp == key_pan) if key_pan else (rn == key_name and ra == key_addr)
        if not hit:
            continue
        sales = _stored_num(ws.cell(row=r, column=P_SALES + 1).value,
                            path, r, "Total Sales")
        purch = _stored_num(ws.cell(row=r, column=P_PURCH + 1).value,
                            path, r, "Total Purchases")
        if is_sale:
            sales = round(sales + total, 2)
        else:
            purch = round(purch + total, 2)
        ws.cell(row=r, column=P_SALES + 1, value=sales)
        ws.cell(row=r, column=P_PURCH + 1, value=purch)
        ws.cell(row=r, column=P_COMBINED + 1, value=round(sales + purch, 2))
        if name.strip():
            ws.cell(row=r, column=P_NAME + 1, value=_safe_text(name.strip()))
        if address.strip():
            ws.cell(row=r, column=P_ADDR + 1, value=_safe_text(address.strip()))
        return

    sales = round(total, 2) if is_sale else 0.0
    purch = 0.0 if is_sale else round(total, 2)
    ws.append([_safe_text(pan.strip()), _safe_text(name.strip()),
               _safe_text(address.strip()), sales, purch,
               round(sales + purch, 2)])


def commit_bill(is_sale: bool, header: dict, lines: list, totals: dict,
                voids: int = None) -> dict:
    """Record one bill and apply it to stock and party totals, atomically.

    Every workbook is prepared in full and staged to a temp file before ANY of
    them is replaced, so a failure part-way through leaves the books exactly as
    they were. Returns {"bill_id", "bill_no"} -- the caller must not
    re-derive the bill number, or the two spellings will drift apart.
    """
    if not lines:
        raise ValueError("A bill needs at least one product line.")

    folder = data_dir()
    ledger = os.path.join(folder, "sales.xlsx" if is_sale else "purchases.xlsx")
    stock_file = os.path.join(folder, "stock.xlsx")
    party_file = os.path.join(folder, "party.xlsx")

    assert_writable(ledger, stock_file, party_file)

    wb_ledger, ws_ledger = _load_rw(ledger, TXN_HEADERS)
    wb_stock, ws_stock = _load_rw(stock_file, STOCK_HEADERS)
    wb_party, ws_party = _load_rw(party_file, PARTY_HEADERS)

    bill_id = next_bill_id(ws_ledger)
    entered = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Bill IDs run per ledger, so the auto number carries S/P too -- otherwise
    # a sale and a purchase could both be called AUTO-00001.
    bill_no = (str(header.get("bill") or "").strip()
               or f"AUTO-{'S' if is_sale else 'P'}{bill_id:05d}")

    for ln in lines:
        ws_ledger.append([
            _safe_text(header.get("date")), _safe_text(bill_no),
            _safe_text(header.get("pan")), _safe_text(header.get("vendor")),
            _safe_text(header.get("address")),
            _safe_text(ln["product"]), ln["qty"], ln["rate"], ln["amount"],
            totals["subtotal"], totals["ecs"], totals["vat_pct"],
            totals["vat_amount"], totals["total"],
            bill_id, entered, voids if voids else None,
        ])
        _apply_stock(ws_stock, stock_file, ln["product"], ln["qty"],
                     add=not is_sale)

    _apply_party(ws_party, party_file, header.get("pan"), header.get("vendor"),
                 header.get("address"), totals["total"], is_sale=is_sale)

    staged = []
    try:
        staged.append((_stage(wb_ledger, ledger), ledger))
        staged.append((_stage(wb_stock, stock_file), stock_file))
        staged.append((_stage(wb_party, party_file), party_file))
    except Exception:
        for tmp, _ in staged:
            _discard(tmp)
        raise
    _commit_staged(staged)

    log.info("committed bill_id=%s (%s) bill_no=%s lines=%d total=%s",
             bill_id, "sale" if is_sale else "purchase", bill_no,
             len(lines), totals["total"])
    return {"bill_id": bill_id, "bill_no": bill_no}


def void_bill(is_sale: bool, bill_id: int) -> int:
    """Cancel a bill by appending a reversing entry, then rebuild the caches.

    History is never rewritten: the original rows stay, and a mirror bill with
    negated quantities and totals is added. Returns the reversing Bill ID.
    """
    folder = data_dir()
    ledger = os.path.join(folder, "sales.xlsx" if is_sale else "purchases.xlsx")
    if bill_id is None:
        raise ValueError(
            "This row has no Bill ID, so the app cannot tell which other rows "
            "belong with it. Cancel it by hand in Excel, then press "
            "Recalculate from Ledgers.")
    target = next((b for b in bills(ledger) if b["bill_id"] == bill_id), None)
    if target is None:
        raise ValueError(f"Bill {bill_id} was not found in this ledger.")
    if target["synthetic"]:
        raise ValueError(f"Bill {bill_id} has no usable Bill ID.")
    if target["voided_by"]:
        raise ValueError(
            f"Bill {bill_id} was already cancelled by bill "
            f"{target['voided_by']}.")
    if target["voids"]:
        raise ValueError(f"Bill {bill_id} is itself a cancellation entry.")

    header = dict(target["header"])
    header["bill"] = f"VOID-{target['header']['bill']}"
    lines = [{"product": ln["product"], "qty": -ln["qty"],
              "rate": ln["rate"], "amount": -ln["amount"]}
             for ln in target["lines"]]
    totals = {k: -v for k, v in target["totals"].items()}
    totals["vat_pct"] = target["totals"]["vat_pct"]      # a rate, not an amount

    new_id = commit_bill(is_sale, header, lines, totals, voids=bill_id)["bill_id"]
    try:
        rebuild_stock()
        rebuild_party_totals()
    except StorageError as exc:
        # The reversal IS recorded. Saying "could not cancel the bill" here
        # would invite the user to do it a second time.
        raise RebuildFailedError(
            f"Bill {bill_id} was cancelled (reversing entry {new_id}), but "
            f"stock and party totals could not be recalculated: {exc}",
            bill_id=new_id) from exc
    log.info("voided bill_id=%s with reversing bill_id=%s", bill_id, new_id)
    return new_id


# --------------------------------------------------------------------------- #
# Rebuilds -- recompute the derived caches from the ledgers
# --------------------------------------------------------------------------- #
def _rewrite(path: str, headers: list, rows: list) -> None:
    """Replace the data rows of a workbook, leaving the user's own tabs alone."""
    if os.path.exists(path):
        wb, ws = _load_rw(path, headers)
        # Deleting the rows would also blank any column the user added beside
        # ours on THIS sheet, leaving an orphaned header. Refuse instead, the
        # same way _migrate refuses a layout it does not recognise.
        extra = [c for c in range(len(headers) + 1, ws.max_column + 1)
                 if any(ws.cell(row=r, column=c).value not in (None, "")
                        for r in range(1, ws.max_row + 1))]
        if extra:
            letters = ", ".join(chr(ord("A") + c - 1) for c in extra)
            raise DataIntegrityError(
                f"{os.path.basename(path)} has your own data in column(s) "
                f"{letters} of the '{SHEET_NAME}' sheet. The app rebuilds this "
                f"sheet from the ledgers and would erase it. Move that data to "
                f"a separate sheet, then try again.")
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        wb = _blank_workbook(headers)
        ws = _sheet(wb, headers)
    for row in rows:
        ws.append(row)
    _save(wb, path)


def migrate_ledgers() -> None:
    """Persist any schema upgrade on the two ledgers, once, at startup.

    Reads deliberately never rewrite the user's file, so a workbook still in
    the old 14-column layout gets its Bill IDs reconstructed in memory on every
    read. Those ids are positional, so re-sorting the sheet in Excel would
    renumber the bills and Void Bill could reverse a different one than the
    dialog named. Writing the upgrade to disk once freezes the identity.
    """
    for path in ledger_files():
        if not os.path.exists(path):
            continue
        try:
            _load_rw(path, TXN_HEADERS)
        except StorageError as exc:
            log.warning("could not upgrade %s: %s", os.path.basename(path), exc)


def rebuild_stock() -> int:
    """Recompute stock.xlsx from every sale and purchase. Returns row count.

    The repair for any drift between the ledgers and the cache, whatever
    caused it. Products are keyed the same way the rest of the module keys
    them, so a name that was previously split across rows is merged back.
    """
    sales, purchases = ledger_files()
    running, display = {}, {}
    for path, sign in ((purchases, 1.0), (sales, -1.0)):
        for r, row in _iter_rows(path, TXN_HEADERS):
            key, text = _key(row[C_PRODUCT]), _display(row[C_PRODUCT])
            if not key:
                continue
            display.setdefault(key, text)
            running[key] = running.get(key, 0.0) + sign * _stored_num(
                row[C_QTY], path, r, "Quantity")

    _rewrite(os.path.join(data_dir(), "stock.xlsx"), STOCK_HEADERS,
             [[_safe_text(display[k]), round(running[k], QTY_DP)]
              for k in sorted(running, key=str.lower)])
    log.info("rebuilt stock.xlsx from the ledgers: %d product(s)", len(running))
    return len(running)


def rebuild_party_totals() -> int:
    """Recompute party.xlsx from every sale and purchase. Returns row count."""
    sales, purchases = ledger_files()
    parties = {}
    for path, field in ((sales, "sales"), (purchases, "purch")):
        # bills() already groups correctly and gives id-less rows their own
        # group, so nothing is double counted and nothing is dropped.
        for bill in bills(path):
            h = bill["header"]
            key_pan, key_name = _key(h["pan"]), _key(h["vendor"])
            key_addr = _key(h["address"])
            if not key_pan and not key_name:
                key_name = "unknown"
            key = ("pan", key_pan) if key_pan else ("na", key_name, key_addr)
            entry = parties.setdefault(key, {
                "pan": _display(h["pan"]),
                "name": _display(h["vendor"]) or "Unknown",
                "address": _display(h["address"]),
                "sales": 0.0, "purch": 0.0})
            entry[field] = round(entry[field] + bill["totals"]["total"], 2)

    _rewrite(os.path.join(data_dir(), "party.xlsx"), PARTY_HEADERS,
             [[_safe_text(e["pan"]), _safe_text(e["name"]), _safe_text(e["address"]),
               e["sales"], e["purch"], round(e["sales"] + e["purch"], 2)]
              for e in parties.values()])
    log.info("rebuilt party.xlsx from the ledgers: %d part(ies)", len(parties))
    return len(parties)


# --------------------------------------------------------------------------- #
# Opening a workbook in Excel
# --------------------------------------------------------------------------- #
def open_file(path: str, headers: list):
    """Open a workbook in the OS default app. Returns (ok, error_message)."""
    try:
        _ensure_file(path, headers)
        system = platform.system()
        if system == "Windows":
            os.startfile(path)  # type: ignore[attr-defined]
        elif system == "Darwin":
            res = subprocess.run(["open", path], timeout=20)
            if res.returncode != 0:
                return False, f"'open' exited with code {res.returncode}."
        else:
            res = subprocess.run(["xdg-open", path], timeout=20)
            if res.returncode != 0:
                return False, f"'xdg-open' exited with code {res.returncode}."
        return True, ""
    except subprocess.TimeoutExpired:
        return False, "The file manager did not respond in time."
    except Exception as exc:  # noqa: BLE001 - surface any launch/IO failure
        log.exception("open_file failed for %s", path)
        return False, str(exc)
